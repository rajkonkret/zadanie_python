import pandas as pd
import random
import os
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill

# 1. Generowanie przykładowych danych (symulacja pliku wejściowego)
def generate_sample_data(filename):
    products = ['Laptop', 'Myszka', 'Klawiatura', 'Monitor', 'Słuchawki', 'Kabel HDMI']
    data = []
    for _ in range(50):
        prod = random.choice(products)
        qty = random.randint(1, 20)
        price = random.randint(20, 4000)
        data.append({'Produkt': prod, 'Ilość': qty, 'Cena_jedn': price})
    
    df = pd.DataFrame(data)
    df.to_excel(filename, index=False)
    print(f"[1/3] Wygenerowano plik źródłowy: {filename}")

# 2. Automatyzacja: Przetwarzanie danych i tworzenie raportu
def process_excel(input_file, output_file):
    if not os.path.exists(input_file):
        print("Brak pliku wejściowego.")
        return

    print("[2/3] Przetwarzanie danych...")
    df = pd.read_excel(input_file)

    # Logika biznesowa: Obliczenie wartości całkowitej
    df['Wartość_całkowita'] = df['Ilość'] * df['Cena_jedn']
    
    # Logika biznesowa: Raport podsumowujący (Pivot)
    summary = df.groupby('Produkt')[['Ilość', 'Wartość_całkowita']].sum().reset_index()
    summary = summary.sort_values(by='Wartość_całkowita', ascending=False)

    # Zapis do nowego pliku z podziałem na arkusze
    with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name='Dane Szczegółowe', index=False)
        summary.to_excel(writer, sheet_name='Raport Podsumowujący', index=False)
    
    print(f"      Zapisano dane do: {output_file}")
    
    # Wywołanie funkcji formatującej
    format_excel_report(output_file)

# 3. Formatowanie (kosmetyka w Excelu przy użyciu openpyxl)
def format_excel_report(filename):
    print("[3/3] Formatowanie stylów w Excelu...")
    wb = load_workbook(filename)
    
    # Formatujemy oba arkusze
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        
        # Styl nagłówka: Pogrubienie, Biały tekst, Niebieskie tło
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill

        # Automatyczne dopasowanie szerokości kolumn
        for col in ws.columns:
            max_length = 0
            column_letter = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            ws.column_dimensions[column_letter].width = max_length + 2

    wb.save(filename)
    print("      Gotowe! Raport jest sformatowany.")

if __name__ == "__main__":
    input_xlsx = "dane_sprzedazowe.xlsx"
    output_xlsx = "raport_finansowy.xlsx"
    
    # Krok 1: Stwórz dane (jeśli nie masz własnych)
    generate_sample_data(input_xlsx)
    
    # Krok 2 i 3: Przetwórz, policz i sformatuj
    process_excel(input_xlsx, output_xlsx)
